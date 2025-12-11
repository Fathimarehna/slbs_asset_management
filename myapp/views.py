from django.shortcuts import render, redirect,get_object_or_404

from django.contrib.auth.decorators import login_required

from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from . models import Asset
from .forms import AssetForm
from .models import Category
from .forms import CategoryForm
from . models import SubCategory
from .forms import SubCategoryForm
from .models import Department
from .forms import DepartmentForm
from .models import Location
from .forms import LocationForm
from django.utils import timezone
from . models import AssetCreate
from .forms import AssetCreateForm
from django.contrib.auth.decorators import user_passes_test

from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Count

# from .forms import UserForm


# Create your views here.


def login_view(request):


    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser(username='admin', password='12345678', email='admin@example.com')
        print("Admin user created")


    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        if username != username.lower():
            messages.error(request, "Invalid username/password")
            return redirect('login')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            

            
            if username == 'admin' and password == '12345678':
                messages.success(request, "Welcome Admin! Login successful.")
                return redirect('admin_dashboard')
            else:
               
                return redirect('user_dashboard')
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'login.html')


def admin_only(user):
    return user.is_superuser

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def admin_dashboard(request):
    total_assets = AssetCreate.objects.count()

    good_assets = AssetCreate.objects.filter(condition='Good').count()
    fair_assets = AssetCreate.objects.filter(condition='Fair').count()
    poor_assets = AssetCreate.objects.filter(condition='Poor').count()

    available_assets = AssetCreate.objects.filter(status=True).count()
    used_assets = AssetCreate.objects.filter(status=False).count()

    context = {
        'total_assets': total_assets,
        'good_assets': good_assets,
        'fair_assets': fair_assets,
        'poor_assets': poor_assets,
        'available_assets': available_assets,
        'used_assets': used_assets,
    }
    return render(request, 'admin_dashboard.html', context)
   

@login_required
def user_dashboard(request):
    return render(request, 'user_dashboard.html')


def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def assetid(request):
    # assets = Asset.objects.all()
    # return render(request, 'assetid.html', {'assets': assets})
    assets = Asset.objects.all()
    form = AssetForm()

    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('assetid')  # refresh page after submission

    return render(request, 'assetid.html', {'assets': assets, 'form': form})




@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def asset_create(request):
    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('assetid')
    else:
        form = AssetForm()
    return render(request, 'asset_form.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def asset_update(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            return redirect('assetid')
    else:
        form = AssetForm(instance=asset)
    return render(request, 'asset_form.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def asset_delete(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        asset.delete()
        return redirect('assetid')
    return render(request, 'assetconfirmdelete.html', {'asset': asset})


@login_required
def toggle_asset_status(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    asset.status = not asset.status  # Toggle between active/inactive
    asset.save()
    return redirect('assetid')

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def category_list(request):
    categories = Category.objects.select_related('asset').all()
    return render(request, 'category_list.html', {'categories': categories})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'category_form.html', {'form': form})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'category_form.html', {'form': form})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'category_confirm_delete.html', {'category': category})

@login_required
def toggle_category_status(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.status = not category.status  # Toggle between active/inactive
    category.save()
    return redirect('category_list')

#subcategory
@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def subcategory_list(request):
    subcategories = SubCategory.objects.all()
    return render(request, 'subcategory.html', {'subcategories': subcategories})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def subcategory_create(request):
    if request.method == 'POST':
        form = SubCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('subcategory_list')
    else:
        form = SubCategoryForm()
    return render(request, 'subcategoryform.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def subcategory_update(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        form =  SubCategoryForm(request.POST, instance=subcategory)
        if form.is_valid():
            form.save()
            return redirect('subcategory_list')
    else:
        form = SubCategoryForm(instance=subcategory)
    return render(request, 'subcategoryform.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def subcategory_delete(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        subcategory.delete()
        return redirect('subcategory_list')
    return render(request, 'subcategory_confirm_delete.html', {'subcategory': subcategory})

@login_required
def toggle_subcategory_status(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    subcategory.status = not subcategory.status  # Toggle between active/inactive
    subcategory.save()
    return redirect('subcategory_list')




#departmentviews
@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def departments_list(request):
    departments=Department.objects.all()
    return render(request, 'departmentslist.html', {'departments': departments})
    
@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def departments_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('departments_list')
    else:
        form = DepartmentForm()
    return render(request, 'departmentsform.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def departments_update(request, pk):
    departments = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form =  DepartmentForm(request.POST, instance=departments)
        if form.is_valid():
            form.save()
            return redirect('departments_list')
    else:
        form = DepartmentForm(instance=departments)
    return render(request, 'departmentsform.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def departments_delete(request, pk):
    departments = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        departments.delete()
        return redirect('departments_list')
    return render(request, 'department_confirm_delete.html', {'departments': departments})

@login_required
def toggle_departments_status(request, pk):
    departments = get_object_or_404(Department, pk=pk)
    departments.status = not departments.status  # Toggle between active/inactive
    departments.save()
    return redirect('departments_list')


#locationsview
@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def locations_list(request):
    locations=Location.objects.all()
    return render(request, 'locations_list.html', {'locations': locations})
    
@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def locations_create(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('locations_list')
    else:
        form = LocationForm()
    return render(request, 'locationsform.html', {'form': form})

@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def locations_update(request, pk):

    locations = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form =  LocationForm(request.POST, instance=locations)
        if form.is_valid():
            form.save()
            return redirect('locations_list')
    else:
        form = LocationForm(instance=locations)
    return render(request, 'locationsform.html', {'form': form})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def locations_delete(request, pk):
    locations = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        locations.delete()
        return redirect('locations_list')
    return render(request, 'location_confirm_delete.html', {'locations': locations})

@login_required
def toggle_locations_status(request, pk):
    locations = get_object_or_404(Location, pk=pk)
    locations.status = not locations.status  # Toggle between active/inactive
    locations.save()
    return redirect('locations_list')




@login_required
def toggle_user_status(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.status = not user.status  # Toggle between active/inactive
    user.save()
    return redirect('user')


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def user_update(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('users_list')
    else:
        form = UserForm(instance=user)
    return render(request, 'user_form.html', {'form': form})



@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def users_view(request):
    users = User.objects.all()
    return render(request, 'users.html', {'users': users})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def add_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        User.objects.create_user(username=username, email=email, password=password)
        return redirect('users')
    return render(request, 'add_user.html')


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        # optional: change password if provided
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
        # update the date
        user.last_login = timezone.now()
        user.save()
        return redirect('users')
    return render(request, 'edit_user.html', {'user': user})


@login_required
@user_passes_test(admin_only, login_url='/userhome/')
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return redirect('users')


@login_required
def assetformlist(request):
    assets=AssetCreate.objects.all()
    return render(request, 'assetlist.html', {'assets': assets})

@login_required
def assetformcreate(request):
    if request.method=='POST':
        print("POST request received")
        print("FILES received:", request.FILES) 
        form=AssetCreateForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            return redirect('assetlist')
    else:
        form=AssetCreateForm()
    return render(request,'assetformcreate.html',{'form':form})

@login_required
def assetform_update(request,pk):
    
    asset = get_object_or_404(AssetCreate, pk=pk)
    if request.method == 'POST':
        form =  AssetCreateForm(request.POST,request.FILES, instance=asset)
        if form.is_valid():
            form.save()
            return redirect('assetlist')
    else:
        form = AssetCreateForm(instance=asset)
    return render(request, 'assetformcreate.html', {'form': form})

@login_required
def assetform_delete(request,pk):
    asset = get_object_or_404(AssetCreate, pk=pk)
    if request.method == 'POST':
       asset.delete()
       return redirect('assetlist') 
    return render(request, 'assetformconfirm_delete.html', {'asset': asset})

@login_required
def assetform_details(request,pk):
    asset=get_object_or_404(AssetCreate,pk=pk)
    return render(request,'assetform_details.html',{'asset':asset})
    
@login_required
def assetformuser(request):
    if request.method=='POST':
        form=AssetCreateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('user_dashboard') 
    else:
        form=AssetCreateForm()
    return render(request,'assetformuser.html',{'form':form})


def load_subcategories(request):
    category_id = request.GET.get('category_id')
    subcategories = SubCategory.objects.filter(category_id=category_id).values('id', 'title')
    return JsonResponse(list(subcategories), safe=False)

@login_required
def add_user(request):
    if request.method == 'POST':
        # userid = request.POST.get('userid')
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if password != confirm_password:
            messages.error(request, "Passwords do not match!")
            return render(request, 'add_user.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists!")
            return render(request, 'add_user.html')

        user = User.objects.create_user(username=username, email=email, password=password)
        user.save()
        messages.success(request, "User created successfully!")
        return redirect('users')

    return render(request, 'add_user.html')



from django.db.models import Q
@user_passes_test(admin_only, login_url='/userhome/')
def asset_report(request):
    assets = AssetCreate.objects.all().order_by("assetname")

    # --- get filter values from form ---

    search = request.GET.get("search", "")
    if search:
        assets = assets.filter(
            Q(assetname__istartswith=search) 
        )

    category = request.GET.get('category')
    subcategory = request.GET.get('subcategory')
    location = request.GET.get('location')
    department = request.GET.get('department')
    condition = request.GET.get('condition')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # --- apply filters dynamically ---
    if category:
        assets = assets.filter(category_id=category)
    if subcategory:
        assets = assets.filter(subcategory=subcategory)
    if location:
        assets = assets.filter(location=location)
    if department:
        assets = assets.filter(department=department)
    if condition:
        assets = assets.filter(condition=condition)
    if date_from:
        assets = assets.filter(purchase_date__gte=date_from)
    if date_to:
        assets = assets.filter(purchase_date__lte=date_to)

    context = {
        'assets': assets,
        'search':search,
        'categories': Category.objects.all(),
        'subcategories':SubCategory.objects.all(),
        'departments': Department.objects.all(),
        'locations': Location.objects.all(),
        'conditions': AssetCreate.CONDITION_CHOICES,
    }
    return render(request, 'report.html', context)

#search option






# AJAX view for dependent subcategory dropdown
def load_subcategories(request):
    category_id = request.GET.get('category_id')
    subcategories = SubCategory.objects.filter(category_id=category_id).values('id', 'title')
    return JsonResponse(list(subcategories), safe=False)



import openpyxl
from django.http import HttpResponse
from .models import AssetCreate

def download_asset_report_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Report"

    # Headers same as UI
    headers = [
        "Asset Name", "Category", "Subcategory", "Department", "Location",
        "Condition", "Purchase Date", "Warranty Expiry", "Assigned To", "Remarks"
    ]
    ws.append(headers)

    assets = AssetCreate.objects.all()

    for a in assets:
        ws.append([
            a.assetname,
            a.category.title if a.category else "",
            a.subcategory.title if a.subcategory else "",
            a.department.title if a.department else "",
            a.location.location if a.location else "",
            a.condition,
            a.purchase_date.strftime("%d-%m-%Y") if a.purchase_date else "",
            a.warrenty_expiry.strftime("%d-%m-%Y") if a.warrenty_expiry else "",
            a.assigned_to,
            a.remarks,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=asset_report.xlsx'

    wb.save(response)
    return response



from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse
from .models import AssetCreate

def download_asset_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=\"asset_report.pdf\"'

    p = canvas.Canvas(response, pagesize=landscape(A4))

    styles = getSampleStyleSheet()
    wrap_style = styles["BodyText"]
    wrap_style.fontSize = 9

    headers = [
        "Asset Name", "Category", "Subcategory", "Department", "Location",
        "Condition", "Purchase Date", "Warranty Expiry", "Assigned To", "Remarks"
    ]

    data = [headers]

    assets = AssetCreate.objects.all()

    for a in assets:
        remarks_para = Paragraph(a.remarks if a.remarks else "", wrap_style)  # WRAP HERE

        data.append([
            a.assetname,
            a.category.title if a.category else "",
            a.subcategory.title if a.subcategory else "",
            a.department.title if a.department else "",
            a.location.location if a.location else "",
            a.condition,
            a.purchase_date.strftime("%d-%m-%Y") if a.purchase_date else "",
            a.warrenty_expiry.strftime("%d-%m-%Y") if a.warrenty_expiry else "",
            a.assigned_to,
            remarks_para,   # USE PARAGRAPH
        ])

    # Create the table with wider remarks column
    table = Table(
        data,
        colWidths=[70, 70, 70, 70, 70, 70, 80, 80, 70, 120]  # Remarks column = 120 width
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]))

    table.wrapOn(p, 30, 400)
    table.drawOn(p, 30, 420)

    p.save()
    return response
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from .models import AssetCreate, Category, SubCategory, Department, Location


def asset_report(request):

    uploaded_data = []     # For Excel preview

    # -------------------------
    # 🔵 HANDLE EXCEL UPLOAD
    # -------------------------
    if request.method == "POST" and request.FILES.get("excel_file"):

        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        sheet = wb.active

        # Store ALL rows for preview (including header)
        for row in sheet.iter_rows(values_only=True):
            uploaded_data.append(row)

        # Insert data into database (skip first row → header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            asset_name, category_name, subcat_name, dept_name, location_name, condition, purchase_date, warr_expiry, assigned_to, remarks = row

            # Fetch related foreign keys
            category = Category.objects.filter(title=category_name).first()
            subcategory = SubCategory.objects.filter(title=subcat_name).first()
            department = Department.objects.filter(title=dept_name).first()
            location = Location.objects.filter(location=location_name).first()

            # Skip if related items not found
            if not category or not subcategory or not department or not location:
                print("⚠️ Skipped row due to missing related data:", row)
                continue

            # Save into DB
            AssetCreate.objects.create(
                assetname=asset_name,
                category=category,
                subcategory=subcategory,
                department=department,
                location=location,
                condition=condition,
                purchase_date=purchase_date,
                warrenty_expiry=warr_expiry,
                assigned_to=assigned_to,
                remarks=remarks,
            )

        # Return SAME PAGE with preview
        return render(request, "report.html", {
            "uploaded_data": uploaded_data,
            "assets": AssetCreate.objects.all(),
            "categories": Category.objects.all(),
            "subcategories": SubCategory.objects.all(),
            "departments": Department.objects.all(),
            "locations": Location.objects.all(),
            "conditions": AssetCreate.CONDITION_CHOICES,
            "search": request.GET.get("search", ""),
        })

    # --------------------------------------
    # 🔵 GET ALL ASSET DATA (PAGE LOAD)
    # --------------------------------------
    assets = AssetCreate.objects.all()

    # Search
    search = request.GET.get("search")
    if search:
        assets = assets.filter(assetname__icontains=search)

    # Filters
    category = request.GET.get("category")
    if category:
        assets = assets.filter(category_id=category)

    subcategory = request.GET.get("subcategory")
    if subcategory:
        assets = assets.filter(subcategory_id=subcategory)

    location = request.GET.get("location")
    if location:
        assets = assets.filter(location_id=location)

    department = request.GET.get("department")
    if department:
        assets = assets.filter(department_id=department)

    condition = request.GET.get("condition")
    if condition:
        assets = assets.filter(condition=condition)

    date_from = request.GET.get("date_from")
    if date_from:
        assets = assets.filter(purchase_date__gte=date_from)

    date_to = request.GET.get("date_to")
    if date_to:
        assets = assets.filter(purchase_date__lte=date_to)

    # Return page normally
    return render(request, "report.html", {
        "assets": assets,
        "categories": Category.objects.all(),
        "subcategories": SubCategory.objects.all(),
        "departments": Department.objects.all(),
        "locations": Location.objects.all(),
        "conditions": AssetCreate.CONDITION_CHOICES,
        "search": search,
    })

